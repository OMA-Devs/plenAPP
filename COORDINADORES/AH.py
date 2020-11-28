#Librería para escribir en los excel.
from openpyxl import load_workbook, Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.styles import Font, Alignment

import re
from copy import copy
import os

path = "\\\\192.168.102.5\\t. de noche\\COORDINADOR AH+\\"

thick_border = Border(right=Side(style='medium'), 
                     top=Side(style='medium'))
top_border = Border(top=Side(style='thin'))

def sortSHEETS(shARR):
	pass

def informePARSER(fName):
	print("Cargando Informe: "+fName)
	origin = load_workbook(path+fName+".xlsx")
	target = Workbook()
	sheet = origin.active
	a = sheet["A"]
	b = sheet["B"]
	ws = 0
	targetSHEET = target.worksheets[ws]
	print(">Tratando el informe:"+fName)
	for ind, cell in enumerate(a):
		if cell.value == None:
			pass
		elif type(cell.value) == str:
			if "DM1" in cell.value:
				tienda = re.search(r'T\d\d\d\d',cell.value).group()
				target.create_sheet(tienda)
				#ws = ws +1 
				targetSHEET = target[tienda]
				targetSHEET.column_dimensions["A"].width = 40
				targetSHEET.column_dimensions["B"].width = 40
				targetSHEET.append(["Reporte Historico de Señales"])
				targetSHEET["A"+str(targetSHEET.max_row)].font = Font(size=24, bold=True)
				targetSHEET.append([])
				targetSHEET.append(["CUENTA: "+cell.value])
				targetSHEET["A"+str(targetSHEET.max_row)].font = Font(bold=True)
				targetSHEET.append([a[ind+1].value])
				targetSHEET["A"+str(targetSHEET.max_row)].font = Font(bold=True)
				targetSHEET.append([""])
				targetSHEET.append([""])
				targetSHEET["B"+str(targetSHEET.max_row)].border = top_border
				targetSHEET["A"+str(targetSHEET.max_row)].border = top_border
				row = ["Fecha y Hora del evento", "Evento"]
				targetSHEET.append(row)
				targetSHEET['A'+str(targetSHEET.max_row)].fill = PatternFill(start_color="000000", end_color="000000", fill_type = "solid") #negro
				targetSHEET['B'+str(targetSHEET.max_row)].fill = PatternFill(start_color="000000", end_color="000000", fill_type = "solid") #negro
				targetSHEET["A"+str(targetSHEET.max_row)].font = Font(bold=True, color="FFFFFF")
				targetSHEET["B"+str(targetSHEET.max_row)].font = Font(bold=True, color="FFFFFF")	
				targetSHEET["A"+str(targetSHEET.max_row)].alignment = Alignment(horizontal="center")
				targetSHEET["B"+str(targetSHEET.max_row)].alignment = Alignment(horizontal="center")
		else:
			try:
				#print(b[ind].value)
				if type(b[ind].value) == str:
					if "SISTEMA ARMADO" in b[ind].value or "SISTEMA DESARMADO" in b[ind].value:
						row = [cell.value, b[ind].value]
						targetSHEET.append(row)
						targetSHEET["B"+str(targetSHEET.max_row)].border = thick_border
						targetSHEET["A"+str(targetSHEET.max_row)].border = thick_border
						if "SISTEMA ARMADO" in b[ind].value:
							targetSHEET['B'+str(targetSHEET.max_row)].fill = PatternFill(start_color="cdb898", end_color="cdb898", fill_type = "solid") #marron
						else:
							targetSHEET['B'+str(targetSHEET.max_row)].fill = PatternFill(start_color="348df9", end_color="348df9", fill_type = "solid") #azul
			except IndexError:
				pass
	target.remove(target["Sheet"])
	###
	final = Workbook()
	sheets = target.worksheets
	actORDER = []
	for sh in sheets:
		actORDER.append(int(sh.title.split("T")[1]))
	actORDER.sort()
	for i in actORDER:
		name = ""
		if len(str(i)) == 3:
			name = "T0"+str(i)
		else:
			name = "T"+str(i)
		try:
			final.create_sheet(name)
			for row in target[name].rows:
				for cell in row:
					new_cell = final[name].cell(row=cell.row, column=cell.col_idx, value= cell.value)
					if cell.has_style:
						new_cell.font = copy(cell.font)
						new_cell.border = copy(cell.border)
						new_cell.fill = copy(cell.fill)
						new_cell.number_format = copy(cell.number_format)
						new_cell.protection = copy(cell.protection)
						new_cell.alignment = copy(cell.alignment)
						new_cell.font = copy(cell.font)
			final[name].column_dimensions["A"].width = 40
			final[name].column_dimensions["B"].width = 40
		except KeyError:
			pass
	final.remove(final["Sheet"])
	try:
		os.remove(path+fName+".xlsx")
		os.remove(path+fName+".xls")
	except FileNotFoundError:
		pass
	final.save(path+fName+".xlsx")

def sorter(wb):
	final = Workbook()
	sheets = wb._sheets
	actORDER = []
	for sh in sheets:
		actORDER.append(int(sh.title.split("T")[1]))
		print(int(sh.title.split("T")[1]))
	newORDER = actORDER.sort()
	for i in newORDER:
		final._sheets.append(target["T"+str(i)])

coordinadores = ["ALBERTO CULEBRAS", "ANGEL AHIJADO", "FRANCISCO DEL AMO", "GARCIA MATEOS", "OSCAR BAZ", "JESUS PERDIGUERO", "RAFAEL TOLDOS"]
#coordinadores = ["JESUS PERDIGUERO"]




for item in coordinadores:
	try:
		informePARSER(item)
	except FileNotFoundError:
		print("INFORME "+item+" NO SE PUEDE USAR POR ALGUNA RAZON")
input("Pulsa Cualquier Tecla para terminar")